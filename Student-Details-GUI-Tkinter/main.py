from tkinter import *
from openpyxl import Workbook, load_workbook
import os

# Create Excel file if not exists
if not os.path.exists("wb.xlsx"):
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Name", "Course", "Semester", "Form No", "Contact", "Email", "Address"])
    wb.save("wb.xlsx")

wb = load_workbook("wb.xlsx")
sheet = wb.active

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert():
    if (
        name_field.get() == "" or
        course_field.get() == "" or
        sem_field.get() == "" or
        form_no_field.get() == "" or
        contact_no_field.get() == "" or
        email_id_field.get() == "" or
        address_field.get() == ""
    ):
        status_label.config(text="⚠ Please fill all fields", fg="red")
    else:
        sheet.append([
            name_field.get(),
            course_field.get(),
            sem_field.get(),
            form_no_field.get(),
            contact_no_field.get(),
            email_id_field.get(),
            address_field.get()
        ])
        wb.save("wb.xlsx")
        clear()
        status_label.config(text="✔ Data saved successfully", fg="green")

root = Tk()
root.title("Student Registration Form")
root.geometry("450x300")
root.resizable(False, False)

Label(root, text="Student Registration Form", font=("Arial", 14, "bold")).grid(row=0, column=1, pady=10)

labels = ["Name", "Course", "Semester", "Form No", "Contact No", "Email ID", "Address"]
entries = []

for i, label in enumerate(labels):
    Label(root, text=label).grid(row=i+1, column=0, padx=10, pady=4, sticky=W)
    entry = Entry(root, width=30)
    entry.grid(row=i+1, column=1)
    entries.append(entry)

name_field, course_field, sem_field, form_no_field, contact_no_field, email_id_field, address_field = entries

Button(root, text="Submit", command=insert, bg="blue", fg="white", width=15).grid(row=9, column=1, pady=10)

status_label = Label(root, text="")
status_label.grid(row=10, column=1)

root.mainloop()